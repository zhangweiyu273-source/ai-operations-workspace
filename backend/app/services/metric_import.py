import csv
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REQUIRED_HEADERS = {"日期", "账号", "内容标题"}
FIELD_MAP = {
    "日期": "metric_date",
    "账号": "account_name",
    "内容标题": "content_title",
    "内容链接": "content_url",
    "内容类型": "content_type",
    "发布时间": "publish_time",
    "曝光": "exposure",
    "播放": "views",
    "点赞": "likes",
    "评论": "comments",
    "收藏": "favorites",
    "分享": "shares",
    "私信": "private_messages",
    "新增线索": "new_leads",
    "有效线索": "valid_leads",
    "高意向": "high_intent_leads",
    "试听": "trial_bookings",
    "成交": "deals",
    "成交金额": "revenue",
    "备注": "notes",
}
INTEGER_FIELDS = {
    "exposure",
    "views",
    "likes",
    "comments",
    "favorites",
    "shares",
    "private_messages",
    "new_leads",
    "valid_leads",
    "high_intent_leads",
    "trial_bookings",
    "deals",
}


def parse_tabular_file(filename: str, content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ValueError("CSV 必须使用 UTF-8 编码") from exc
        reader = csv.DictReader(StringIO(text))
        return list(reader.fieldnames or []), list(reader)
    if suffix == ".xlsx":
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(values, ())]
        rows = [
            dict(zip(headers, row, strict=False))
            for row in values
            if any(value is not None and str(value).strip() for value in row)
        ]
        workbook.close()
        return headers, rows
    raise ValueError("仅支持 .csv 和 .xlsx 文件")


def normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    data: dict[str, Any] = {}
    for header, field in FIELD_MAP.items():
        value = row.get(header)
        if isinstance(value, str):
            value = value.strip()
        data[field] = value if value not in (None, "") else None
    raw_date = data["metric_date"]
    if isinstance(raw_date, datetime):
        data["metric_date"] = raw_date.date()
    elif isinstance(raw_date, date):
        pass
    elif raw_date:
        try:
            data["metric_date"] = date.fromisoformat(str(raw_date))
        except ValueError as exc:
            raise ValueError("日期必须为 YYYY-MM-DD") from exc
    for field in INTEGER_FIELDS:
        value = data[field]
        if value is None:
            data[field] = 0
            continue
        try:
            decimal = Decimal(str(value))
            if decimal != decimal.to_integral_value() or decimal < 0:
                raise ValueError
            data[field] = int(decimal)
        except (InvalidOperation, ValueError) as exc:
            raise ValueError(f"{field} 必须是非负整数") from exc
    try:
        data["revenue"] = Decimal(str(data["revenue"] or 0)).quantize(Decimal("0.01"))
        if data["revenue"] < 0:
            raise ValueError
    except (InvalidOperation, ValueError) as exc:
        raise ValueError("revenue 必须是非负金额") from exc
    if data["publish_time"] and not isinstance(data["publish_time"], datetime):
        try:
            data["publish_time"] = datetime.fromisoformat(str(data["publish_time"]))
        except ValueError as exc:
            raise ValueError("发布时间格式不合法") from exc
    return data
